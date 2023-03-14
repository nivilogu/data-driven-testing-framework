# excel_functions.py

# functional file for Excel operations

from openpyxl import load_workbook


class Suman_Excel_Functions:
    def __init__(self, excel_file_name, sheet_name):
        self.file = excel_file_name
        self.sheet = sheet_name

    # fetch the row count of the excel file
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.max_row)

    # fetch the column count of the excel file
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.max_column)

    # read the data from the Excel file
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        data = sheet.cell(row=row_number, column=column_number).value
        return data

    # write the data to the Excel file
    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.file)
